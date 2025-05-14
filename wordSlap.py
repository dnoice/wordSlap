#!/usr/bin/env python3

"""
Title: WORD SLAP ULTRA 👋💥
Author: Not You!

Description: The ULTIMATE word analysis script that slaps so hard it'll leave your vocabulary in tears!

This isn't just word analysis - this is a VERBAL MASSACRE featuring:
✓ WordNet POS tagging that will make linguists question their careers
✓ Sentiment analysis so accurate it can detect your ex's passive-aggressive texts
✓ Visualizations so stunning they belong in the MOMA
✓ More export formats than you have dating apps
✓ Statistics so advanced they required their own mathematics degree
✓ Error handling that catches problems before they even happen
✓ Performance that makes quantum computers look like potatoes

Don't just analyze words. DOMINATE THEM.
"""

import json
import csv
import yaml
import nltk
import os
import time
import logging
import threading
import shutil
import random
import sys
import textwrap
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
from typing import Dict, List, Tuple, Optional, Any, Set
from nltk.corpus import wordnet as wn
from nltk.corpus import sentiwordnet as swn
from nltk.sentiment import SentimentIntensityAnalyzer
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer
from operator import itemgetter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.ticker import MaxNLocator
import numpy as np
from functools import lru_cache
import seaborn as sns
from tqdm import tqdm
import colorama
from colorama import Fore, Back, Style
from tabulate import tabulate
import concurrent.futures
import markdown
import psutil
import platform


# ==================================================================================
# CONFIGURATION - Settings so flexible they should join the Olympic gymnastics team
# ==================================================================================
 DDsddsdddsessessae\TV:@@æq¹...2²r¹1¹111¹1111111¹¹

# Initialize colorama for cross-platform colored terminal output
colorama.init()


# File paths - Because your words deserve a proper home
INPUT_FILE = "words.txt"
OUTPUT_DIR = "output"
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_SUBDIR = os.path.join(OUTPUT_DIR, f"analysis_{TIMESTAMP}")
JSON_OUTPUT = os.path.join(OUTPUT_SUBDIR, "wordnet_stats.json")
YAML_OUTPUT = os.path.join(OUTPUT_SUBDIR, "wordnet_stats.yaml")
CSV_OUTPUT = os.path.join(OUTPUT_SUBDIR, "wordnet_words_by_type.csv")
XLSX_OUTPUT = os.path.join(OUTPUT_SUBDIR, "wordnet_words_by_type.xlsx")
TEXT_OUTPUT = os.path.join(OUTPUT_SUBDIR, "word_analysis_report.txt")
HTML_OUTPUT = os.path.join(OUTPUT_SUBDIR, "word_analysis_report.html")
MARKDOWN_OUTPUT = os.path.join(OUTPUT_SUBDIR, "word_analysis_report.md")
LOG_FILE = os.path.join(OUTPUT_SUBDIR, "analyzer.log")
CHART_DIR = os.path.join(OUTPUT_SUBDIR, "charts")
MAIN_CHART_OUTPUT = os.path.join(CHART_DIR, "word_distribution.png")
POS_CHART_OUTPUT = os.path.join(CHART_DIR, "pos_distribution.png")
LENGTH_CHART_OUTPUT = os.path.join(CHART_DIR, "length_distribution.png")
SENTIMENT_CHART_OUTPUT = os.path.join(CHART_DIR, "sentiment_distribution.png")
FREQ_CHART_OUTPUT = os.path.join(CHART_DIR, "frequency_distribution.png")
CLUSTER_CHART_OUTPUT = os.path.join(CHART_DIR, "word_clusters.png")
STATS_DUMP_FILE = os.path.join(OUTPUT_SUBDIR, "deep_stats.json")
CACHE_DIR = os.path.join(OUTPUT_DIR, "cache")
CACHE_FILE = os.path.join(CACHE_DIR, "wordnet_cache.json")


# Word categories - Because labels are for chumps
CATEGORIES = {
    "n": "Noun",
    "v": "Verb",
    "a": "Adjective",
    "r": "Adverb",
    "s": "Adjective Satellite",  # WordNet uses 's' for adjective satellites
}


# Visualization settings - Colors so vibrant they'll make your retinas high-five
COLORS = {
    "Noun": "#4287f5",  # Blue
    "Verb": "#f542a7",  # Pink
    "Adjective": "#42f56f",  # Green
    "Adjective Satellite": "#42d1f5",  # Cyan
    "Adverb": "#f5d442",  # Yellow
    "Unknown": "#a6a6a6",  # Gray
    "Positive": "#42f56f",  # Green for positive sentiment
    "Negative": "#f54242",  # Red for negative sentiment
    "Neutral": "#4287f5",  # Blue for neutral sentiment
}


# Expanded color palette for more visually stunning charts
EXPANDED_COLORS = list(mcolors.TABLEAU_COLORS.values()) + list(
    mcolors.CSS4_COLORS.keys()
)


# Performance settings - We're not just fast, we're "why-is-my-hair-blown-back" fast
CACHE_SIZE = 5000  # Number of words to cache for WordNet lookups
BATCH_SIZE = 1000  # Batch size for processing large files
MAX_WORKERS = max(4, os.cpu_count())  # Maximum number of worker threads
PROGRESS_INTERVAL = 100  # How often to show progress (every N words)
USE_PARALLEL = True  # Use parallel processing for heavy operations
USE_CACHE = True  # Use cache for repeated operations


# Threshold values - We have standards, unlike your ex
MIN_WORD_LENGTH = 1  # Minimum word length to process
MAX_WORD_LENGTH = 45  # Maximum word length to process
MIN_CONFIDENCE = 0.3  # Minimum confidence for classification
SENTIMENT_THRESHOLD = 0.2  # Threshold for determining sentiment
RARE_WORD_THRESHOLD = 5  # Threshold for considering a word "rare"
PREMIUM_WORD_LENGTH = 8  # Length at which a word becomes "premium grade"


# Reporting options - Because if you're gonna flex, flex HARD
REPORT_TOP_N = 15  # Number of top words to show in each category
VERBOSE_OUTPUT = True  # Show detailed console output
USE_FANCY_PROGRESS = True  # Use fancy progress bars with tqdm
ENABLE_SOUNDS = False  # Enable sound effects (for terminals that support it)
USE_COLORS = True  # Use colors in terminal output
EASTER_EGGS = True  # Enable Easter eggs and random sassy comments
ENABLE_MEMORY_TRACKING = True  # Track memory usage


# Analysis options - The kind of extra that makes your buddy weep tears of envy
ENABLE_SENTIMENT = True  # Enable sentiment analysis
ENABLE_FREQUENCY = True  # Enable word frequency analysis
ENABLE_CLUSTERING = True  # Enable word clustering
ENABLE_SYLLABLE_COUNT = True  # Enable syllable counting
ENABLE_READABILITY = True  # Enable readability metrics
ENABLE_LINGUISTIC_FEATURES = True  # Enable additional linguistic features
ENABLE_ETYMOLOGY = False  # Enable etymology lookup (requires internet)


# ==================================================================================
# SETUP - Prepare for liftoff, mortals
# ==================================================================================


# Configure logging - Because when we fail, we fail SPECTACULARLY
def setup_logging():
    """Set up logging configuration with more sass than a teenager."""

    os.makedirs(OUTPUT_SUBDIR, exist_ok=True)

    # Define custom log formatter with colors
    class ColoredFormatter(logging.Formatter):
        """Logging formatter with colors that pop harder than boy bands."""

        COLORS = {
            "DEBUG": Fore.CYAN,
            "INFO": Fore.GREEN,
            "WARNING": Fore.YELLOW,
            "ERROR": Fore.RED,
            "CRITICAL": Fore.RED + Back.WHITE + Style.BRIGHT,
        }

        def format(self, record):
            levelname = record.levelname
            msg = super().format(record)
            if USE_COLORS and levelname in self.COLORS:
                return f"{self.COLORS[levelname]}{msg}{Style.RESET_ALL}"
            return msg

    # Create formatters
    file_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    console_formatter = ColoredFormatter("%(asctime)s - %(levelname)s - %(message)s")

    # Set up file handler
    file_handler = logging.FileHandler(LOG_FILE)
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(file_formatter)

    # Set up console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(console_formatter)

    # Configure root logger
    logging.basicConfig(level=logging.DEBUG, handlers=[file_handler, console_handler])

    # Add some attitude
    logging.info(
        f"WORD SLAP ULTRA is initializing. Prepare your dictionaries for ANNIHILATION."
    )
    logging.debug(f"Output directory: {OUTPUT_SUBDIR}")
    logging.debug(f"Log file: {LOG_FILE}")


# First-time NLTK setup - Because we don't half-ass our NLP
def setup_nltk():
    """Download required NLTK resources if not already present.
    More resources = more POWER."""

    resources = [
        ("wordnet", "For basic POS tagging"),
        ("sentiwordnet", "For in-depth sentiment analysis"),
        ("vader_lexicon", "For advanced sentiment analysis"),
        ("punkt", "For tokenization so precise it could split atoms"),
        ("averaged_perceptron_tagger", "For POS tagging that makes linguists jealous"),
    ]

    logging.info("Checking NLTK resources (because we're THOROUGH like that)...")

    for resource, desc in resources:
        try:
            nltk.data.find(
                f"corpora/{resource}"
                if "wordnet" in resource or "lexicon" in resource
                else f"tokenizers/{resource}"
            )
            logging.debug(f"{resource} already downloaded")
        except LookupError:
            logging.info(f"Downloading {resource} ({desc})...")
            nltk.download(resource, quiet=True)
            logging.info(f"{resource} downloaded successfully")


# Create required directories
def setup_directories():
    """Create all required directories with the confidence of a cat walking on a keyboard."""

    dirs = [OUTPUT_DIR, OUTPUT_SUBDIR, CHART_DIR, CACHE_DIR]
    for directory in dirs:
        os.makedirs(directory, exist_ok=True)
        logging.debug(f"Directory created or verified: {directory}")


# Display fancy ASCII art banner because we're EXTRA like that
def display_banner():
    """Display a ridiculously over-the-top banner because subtlety is for the weak."""

    banner = r"""
 _       _____  ___   ___  ______      ______  __      ___   ___     __  __ __    ________  ___   ___  
| |     / / _ \/ _ | / _ \/ __/ /     / __/ / / /     / _ | / _ \   / / / //_/___/_  __/ / / _ | / _ \ 
| | /| / / , _/ __ |/ // /\ \/ /____ _\ \/ /_/ /___  / __ |/ ___/  / /_/ / /___/  / / / /_/ __ |/ , _/ 
|_|/_|/_/_/|_/_/ |_/____/___/____/(_)___/\____/___/ /_/ |_/_/     /____/_/       /_/ /___/_/ |_/_/|_|  
                                                                                                      
ANALYZING WORDS SO HARD EVEN YOUR DICTIONARY IS INTIMIDATED
    """

    if USE_COLORS:
        # Choose a random color for the banner each time because WHY NOT?
        colors = [Fore.CYAN, Fore.MAGENTA, Fore.YELLOW, Fore.GREEN, Fore.RED]
        banner_color = random.choice(colors)
        print(f"{banner_color}{banner}{Style.RESET_ALL}")
    else:
        print(banner)

    # Add a sassy tagline
    taglines = [
        "Your words are about to get SLAPPED into next Tuesday!",
        "Prepare for word analysis so thorough it borders on stalking!",
        "We don't just analyze words, we JUDGE them!",
        "Making other word analyzers look like kindergarten projects since 2022!",
        "If Shakespeare were alive, he'd be using THIS tool!",
        "Not your grampa's word analyzer... unless your grampa is AWESOME!",
        "Words fear us. Linguistics departments want to BE us.",
        "The only word analyzer with a black belt in semantic destruction!",
    ]

    print(f"\n{Fore.CYAN}{random.choice(taglines)}{Style.RESET_ALL}\n")


def play_sound(sound_type="start"):
    """Play a sound effect if enabled and system supports it."""

    if not ENABLE_SOUNDS:
        return

    # This is just a placeholder. In a real implementation, you might use
    # a library like playsound, pygame, or even just the system's beep.
    sounds = {
        "start": "\a",  # System beep
        "complete": "\a\a",  # Double beep
        "error": "\a\a\a",  # Triple beep
    }

    try:
        sys.stdout.write(sounds.get(sound_type, "\a"))
        sys.stdout.flush()
    except:
        pass  # Silently fail if sounds aren't supported


# ==================================================================================
# CACHE MANAGEMENT - Because we're efficient AND lazy
# ==================================================================================

def load_cache():
    """Load word information from cache file."""

    if not USE_CACHE:
        return {}

    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                logging.info(
                    f"Loading cache from {CACHE_FILE}... (working smarter, not harder)"
                )
                return json.load(f)
    except Exception as e:
        logging.warning(f"Error loading cache: {e}")

    return {}


def save_cache(cache_data):
    """Save word information to cache file for future runs."""

    if not USE_CACHE:
        return

    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            logging.info(
                f"Saving cache to {CACHE_FILE}... (future you will thank present you)"
            )
            json.dump(cache_data, f)
    except Exception as e:
        logging.warning(f"Error saving cache: {e}")


# ==================================================================================
# WORD CLASSIFICATION - Where words go to be judged, HARSHLY
# ==================================================================================


@lru_cache(maxsize=CACHE_SIZE)
def get_synsets(word):
    """Get WordNet synsets for a word with caching for performance.
    We're fast AND smart, unlike your last relationship."""

    return wn.synsets(word)


@lru_cache(maxsize=CACHE_SIZE)
def get_sentiment(word):
    """Get sentiment analysis for a word.
    We know how you FEEL about words before you do."""

    if not ENABLE_SENTIMENT:
        return {"positive": 0.0, "negative": 0.0, "neutral": 1.0, "compound": 0.0}

    try:
        # Get SentiWordNet sentiment
        synsets = get_synsets(word)

        if not synsets:
            # Fall back to VADER for words not in WordNet
            analyzer = SentimentIntensityAnalyzer()
            scores = analyzer.polarity_scores(word)
            return {
                "positive": scores["pos"],
                "negative": scores["neg"],
                "neutral": scores["neu"],
                "compound": scores["compound"],
            }

        # Calculate sentiment from SentiWordNet
        pos_score = 0.0
        neg_score = 0.0
        count = 0

        for synset in synsets:
            try:
                swn_synset = swn.senti_synset(synset.name())
                pos_score += swn_synset.pos_score()
                neg_score += swn_synset.neg_score()
                count += 1
            except:
                continue

        if count > 0:
            pos_score /= count
            neg_score /= count

        neutral_score = 1.0 - (pos_score + neg_score)
        compound = pos_score - neg_score

        return {
            "positive": pos_score,
            "negative": neg_score,
            "neutral": neutral_score,
            "compound": compound,
        }

    except Exception as e:
        logging.debug(f"Error getting sentiment for '{word}': {e}")
        return {"positive": 0.0, "negative": 0.0, "neutral": 1.0, "compound": 0.0}


def calculate_confidence(frequencies):
    """Calculate confidence score based on POS frequencies.
    Our confidence is as unshakable as a rock... in cement... on bedrock."""

    if not frequencies:
        return 0.0

    total = sum(frequencies.values())
    highest = max(frequencies.values())
    return highest / total


@lru_cache(maxsize=CACHE_SIZE)
def count_syllables(word):
    """Count syllables in a word using a basic heuristic approach.
    More accurate than your middle school English teacher."""

    if not ENABLE_SYLLABLE_COUNT:
        return 1

    word = word.lower()

    # Remove non-alpha characters
    word = "".join(c for c in word if c.isalpha())

    if not word:
        return 0

    # Exception cases
    exceptions = {
        "area": 3,
        "aria": 3,
        "aura": 2,
        "every": 3,
        "evening": 3,
        "everything": 4,
        "focusing": 3,
        "seventh": 2,
        "create": 2,
        "creature": 2,
        "credit": 2,
    }

    if word in exceptions:
        return exceptions[word]

    # Count vowel groups
    vowels = "aeiouy"
    prev_is_vowel = False
    count = 0

    # Check each character
    for char in word:
        is_vowel = char in vowels
        if is_vowel and not prev_is_vowel:
            count += 1
        prev_is_vowel = is_vowel

    # Handle special cases
    if word.endswith("e"):
        count -= 1
    if word.endswith("le") and len(word) > 2 and word[-3] not in vowels:
        count += 1
    if count == 0:
        count = 1

    return count


def get_readability_level(avg_syllables, avg_word_length):
    """Get an approximate readability level based on word metrics.
    We judge your vocabulary like it's a reality TV show contestant."""

    if not ENABLE_READABILITY:
        return "Unknown"

    # Very basic approximation based on syllables and word length
    score = (avg_syllables * 0.39) + (avg_word_length * 0.05) - 2.0

    if score <= 1:
        return "Elementary"
    elif score <= 3:
        return "Easy"
    elif score <= 5:
        return "Average"
    elif score <= 7:
        return "Difficult"
    else:
        return "Advanced"


def get_linguistic_features(word):
    """Get additional linguistic features for a word.
    We analyze words like they're suspects in a crime documentary."""

    if not ENABLE_LINGUISTIC_FEATURES:
        return {}

    features = {
        "palindrome": word.lower() == word.lower()[::-1],
        "alliteration": False,
        "contains_prefix": False,
        "contains_suffix": False,
        "rare_letter": False,
        "euphony": False,
    }

    # Check for common prefixes
    prefixes = [
        "un",
        "re",
        "in",
        "im",
        "dis",
        "en",
        "em",
        "non",
        "pre",
        "anti",
        "auto",
        "bi",
        "co",
    ]

    for prefix in prefixes:
        if word.lower().startswith(prefix):
            features["contains_prefix"] = True
            break

    # Check for common suffixes
    suffixes = [
        "ing",
        "ed",
        "ly",
        "er",
        "est",
        "ment",
        "ness",
        "ful",
        "less",
        "able",
        "ible",
        "al",
        "ial",
    ]

    for suffix in suffixes:
        if word.lower().endswith(suffix):
            features["contains_suffix"] = True
            break

    # Check for rare letters
    rare_letters = "jkqxz"
    features["rare_letter"] = any(letter in rare_letters for letter in word.lower())

    # Check for euphony (pleasant sounding words often have balanced vowels/consonants)
    vowels = sum(1 for char in word.lower() if char in "aeiou")
    consonants = sum(
        1 for char in word.lower() if char.isalpha() and char not in "aeiou"
    )

    # Words with balanced vowels and consonants tend to sound pleasant
    vowel_ratio = vowels / len(word) if len(word) > 0 else 0
    features["euphony"] = 0.3 <= vowel_ratio <= 0.6

    # Check for alliteration potential (repeated initial consonant sounds)
    if len(word) >= 2:
        features["alliteration"] = word[0] in "bcdfghjklmnpqrstvwxyz"

    return features


def classify_wordnet(word, word_cache=None):
    """
    Classify a word using WordNet and return its most likely part of speech.
    Our classifications are more accurate than your horoscope.

    Args:
        word: The word to classify
        word_cache: Optional cache for storing results

    Returns:
        tuple: (most_common_pos, pos_frequencies, confidence_score, additional_info)
    """

    # Check cache first if provided
    if word_cache is not None and word in word_cache:
        return word_cache[word]

    if not word or len(word) < MIN_WORD_LENGTH or len(word) > MAX_WORD_LENGTH:
        return "Unknown", {}, 0.0, {}

    # Get synsets from WordNet
    synsets = get_synsets(word)

    if not synsets:
        return "Unknown", {}, 0.0, {}

    # Count frequency of each part of speech
    pos_freq = Counter()

    for syn in synsets:
        pos = syn.pos()
        if pos in CATEGORIES:
            pos_freq[CATEGORIES[pos]] += 1

    if not pos_freq:
        return "Unknown", {}, 0.0, {}

    # Find most common part of speech
    most_common, count = pos_freq.most_common(1)[0]
    confidence = calculate_confidence(pos_freq)

    # Get additional word information
    additional_info = {
        "sentiment": get_sentiment(word),
        "syllables": count_syllables(word),
        "linguistic_features": get_linguistic_features(word),
    }

    # If confidence is too low, mark as ambiguous
    if confidence < MIN_CONFIDENCE:
        most_common = f"{most_common} (ambiguous)"

    # Cache the result if we have a cache
    if word_cache is not None:
        word_cache[word] = (most_common, dict(pos_freq), confidence, additional_info)

    return most_common, dict(pos_freq), confidence, additional_info


# ==================================================================================
# DATA PROCESSING - Where magic happens and your buddy's analyzer weeps in shame
# ==================================================================================


def read_words(path):
    """Read words from input file with error handling.
    We handle files like a boss, not like Windows Explorer."""

    words = []

    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                word = line.strip()
                if word:  # Skip empty lines
                    words.append(word)

        # Add some sass
        word_count = len(words)

        if word_count == 0:
            logging.error(f"The file {path} is as empty as my ex's promises")
        elif word_count < 10:
            logging.info(
                f"Read {word_count} words from {path}. That's it? My grocery list is longer!"
            )
        elif word_count < 100:
            logging.info(
                f"Read {word_count} words from {path}. I've seen bigger lists in fortune cookies!"
            )
        elif word_count < 1000:
            logging.info(
                f"Read {word_count} words from {path}. Not bad, but I've seen dictionaries with more personality!"
            )
        else:
            logging.info(
                f"Read {word_count} words from {path}. NOW we're talking! Time to flex on these words!"
            )

        return words

    except FileNotFoundError:
        logging.error(f"Input file not found: {path}. Did you forget how to type? 🙄")
        return []

    except Exception as e:
        logging.error(
            f"Error reading input file: {e}. Not my fault, but I'll fix YOUR mistake!"
        )
        return []


def process_batch(batch, word_cache):
    """
    Process a batch of words for parallel execution.
    Multitasking like your mom telling you to clean your room while cooking dinner.
    """

    results = {}

    for word in batch:
        most_common, freq_map, confidence, additional_info = classify_wordnet(
            word, word_cache
        )

        results[word] = {
            "main_type": most_common.split(" ")[0]
            if most_common
            else "Unknown",  # Remove "(ambiguous)" part
            "pos_frequencies": freq_map,
            "confidence": confidence,
            "length": len(word),
            "synset_count": len(get_synsets(word)),
            "sentiment": additional_info.get("sentiment", {}),
            "syllables": additional_info.get("syllables", 1),
            "linguistic_features": additional_info.get("linguistic_features", {}),
        }

    return results

def analyze_words(words):
    """

    Analyze words to collect statistics and categorize them.

    This function puts the THOR in THORough.



    Args:

        words: List of words to analyze



    Returns:

        tuple: (stats, detailed, word_info, metrics)

    """

    # Load existing cache if enabled

    word_cache = load_cache() if USE_CACHE else {}

    stats = defaultdict(int)

    detailed = defaultdict(list)

    word_info = {}

    # Additional metrics - Because basic stats are for BASIC analyzers

    metrics = {
        "total_words": len(words),
        "unique_words": 0,
        "avg_word_length": 0,
        "avg_confidence": 0,
        "avg_syllables": 0,
        "processing_time": 0,
        "ambiguous_words": 0,
        "unknown_words": 0,
        "positive_words": 0,
        "negative_words": 0,
        "neutral_words": 0,
        "palindromes": 0,
        "rare_letter_words": 0,
        "prefix_words": 0,
        "suffix_words": 0,
        "euphonious_words": 0,
        "alliteration_words": 0,
        "length_distribution": defaultdict(int),
        "confidence_distribution": defaultdict(int),
        "syllable_distribution": defaultdict(int),
        "premium_words": 0,  # Words above a certain length
        "readability_index": 0,
        "memory_usage": 0,
    }

    if ENABLE_MEMORY_TRACKING:
        # Get baseline memory usage

        metrics["initial_memory"] = psutil.Process().memory_info().rss / (
            1024 * 1024
        )  # In MB

    # Track unique words

    unique_words = set()

    total_length = 0

    total_confidence = 0

    total_syllables = 0

    start_time = time.time()

    # Create batches for parallel processing

    word_batches = [words[i : i + BATCH_SIZE] for i in range(0, len(words), BATCH_SIZE)]

    # Process words with fancy progress bar

    if USE_FANCY_PROGRESS:
        pbar = tqdm(
            total=len(words),
            desc="Analyzing words",
            unit="words",
            bar_format="{l_bar}%s{bar}%s{r_bar}" % (Fore.GREEN, Style.RESET_ALL)
            if USE_COLORS
            else None,
        )

    if USE_PARALLEL and len(words) > BATCH_SIZE:
        # Parallel processing for large datasets

        logging.info(
            f"Unleashing {MAX_WORKERS} worker threads to conquer your words in parallel!"
        )

        batch_results = {}

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            # Submit batch processing tasks

            future_to_batch = {
                executor.submit(process_batch, batch, word_cache): i
                for i, batch in enumerate(word_batches)
            }

            # Process completed batches

            for future in concurrent.futures.as_completed(future_to_batch):
                batch_idx = future_to_batch[future]

                try:
                    result = future.result()

                    batch_results.update(result)

                    if USE_FANCY_PROGRESS:
                        pbar.update(len(word_batches[batch_idx]))

                except Exception as e:
                    logging.error(f"Error processing batch {batch_idx}: {e}")

        # Use batch results

        word_info = batch_results

    else:
        # Sequential processing for smaller datasets

        for i, word in enumerate(words):
            # Show progress for large files

            if USE_FANCY_PROGRESS:
                pbar.update(1)

            elif i % PROGRESS_INTERVAL == 0 and i > 0 and VERBOSE_OUTPUT:
                logging.info(
                    f"Processing... {i}/{len(words)} words ({(i/len(words))*100:.1f}%)"
                )

            # Skip duplicates but count them

            if word in unique_words:
                continue

            unique_words.add(word)

            # Classify the word

            main_type, freq_map, confidence, additional_info = classify_wordnet(
                word, word_cache
            )

            # Update statistics

            category = main_type.split(" ")[0]  # Remove "(ambiguous)" part if present

            stats[category] += 1

            detailed[category].append(word)

            # Update metrics

            total_length += len(word)

            total_confidence += confidence

            syllables = additional_info.get("syllables", 1)

            total_syllables += syllables

            metrics["length_distribution"][len(word)] += 1

            metrics["confidence_distribution"][round(confidence, 1)] += 1

            metrics["syllable_distribution"][syllables] += 1

            # Check for "premium" words (longer than threshold)

            if len(word) >= PREMIUM_WORD_LENGTH:
                metrics["premium_words"] += 1

            if "ambiguous" in main_type:
                metrics["ambiguous_words"] += 1

            if category == "Unknown":
                metrics["unknown_words"] += 1

            # Track sentiment

            sentiment = additional_info.get("sentiment", {})

            compound = sentiment.get("compound", 0)

            if compound >= SENTIMENT_THRESHOLD:
                metrics["positive_words"] += 1

            elif compound <= -SENTIMENT_THRESHOLD:
                metrics["negative_words"] += 1

            else:
                metrics["neutral_words"] += 1

            # Track linguistic features

            ling_features = additional_info.get("linguistic_features", {})

            if ling_features.get("palindrome", False):
                metrics["palindromes"] += 1

            if ling_features.get("rare_letter", False):
                metrics["rare_letter_words"] += 1

            if ling_features.get("contains_prefix", False):
                metrics["prefix_words"] += 1

            if ling_features.get("contains_suffix", False):
                metrics["suffix_words"] += 1

            if ling_features.get("euphony", False):
                metrics["euphonious_words"] += 1

            if ling_features.get("alliteration", False):
                metrics["alliteration_words"] += 1

            # Store word information

            word_info[word] = {
                "main_type": category,
                "pos_frequencies": freq_map,
                "confidence": confidence,
                "length": len(word),
                "synset_count": len(get_synsets(word)),
                "sentiment": sentiment,
                "syllables": syllables,
                "linguistic_features": ling_features,
            }

    if USE_FANCY_PROGRESS:
        pbar.close()

    # Save the cache for future runs

    if USE_CACHE:
        save_cache(word_cache)

    # Calculate final metrics

    metrics["unique_words"] = len(unique_words)

    if metrics["unique_words"] > 0:
        metrics["avg_word_length"] = total_length / metrics["unique_words"]

        metrics["avg_confidence"] = total_confidence / metrics["unique_words"]

        metrics["avg_syllables"] = total_syllables / metrics["unique_words"]

    # Calculate readability

    metrics["readability_index"] = get_readability_level(
        metrics["avg_syllables"], metrics["avg_word_length"]
    )

    metrics["processing_time"] = time.time() - start_time

    if ENABLE_MEMORY_TRACKING:
        # Get final memory usage

        metrics["final_memory"] = psutil.Process().memory_info().rss / (
            1024 * 1024
        )  # In MB

        metrics["memory_usage"] = metrics["final_memory"] - metrics["initial_memory"]

    return stats, detailed, word_info, metrics


# ==================================================================================

# REPORTING AND VISUALIZATION - Where we make your buddy's charts look like crayon drawings

# ==================================================================================


def get_sassy_comment(metric_name, value):
    """Generate sassy comments based on metrics.

    Because numbers without attitude are just... numbers."""

    comments = {
        "total_words": [
            f"That's {value} more words than most people know!",
            f"{value} words? I've seen dictionaries smaller than that!",
            f"{value} words? That's like a novel! Well, maybe a pamphlet.",
            f"Processed {value} words faster than you can say 'supercalifragilisticexpialidocious'!",
        ],
        "unique_words": [
            f"{value} unique words - more variety than your dating history!",
            f"{value} unique words - that's more than most politicians use in their entire careers!",
            f"Found {value} unique words. Shakespeare is getting jealous!",
            f"{value} unique words? Your vocabulary is showing off!",
        ],
        "avg_word_length": [
            f"Average length of {value:.2f} - compensating for something?",
            f"{value:.2f} letters per word. Brevity is NOT your strong suit!",
            f"Words averaging {value:.2f} letters. Trying to win at Scrabble?",
            f"Average length: {value:.2f} - sized for maximum tongue-twisting!",
        ],
        "readability_index": [
            f"Readability level: {value}. Even your cat could understand these words!",
            f"Readability level: {value}. Shakespeare would be... confused.",
            f"Readability level: {value}. Perfect for your next academic paper!",
            f"Readability level: {value}. Your words are as accessible as Fort Knox!",
        ],
        "processing_time": [
            f"Processed in {value:.2f} seconds - Flash who?",
            f"Analysis completed in {value:.2f} seconds - while you were still blinking!",
            f"Took {value:.2f} seconds - quantum computers are calling for tips!",
            f"Processing time: {value:.2f} seconds. I've had slower coffee breaks!",
        ],
        "memory_usage": [
            f"Used {value:.2f}MB of memory - your words aren't THAT important!",
            f"Memory used: {value:.2f}MB - cheaper than therapy!",
            f"Memory usage: {value:.2f}MB - I've stored shopping lists that took more space!",
            f"Consumed {value:.2f}MB of precious memory. Hope these words were worth it!",
        ],
    }

    if metric_name in comments:
        return random.choice(comments[metric_name])

    return ""


def generate_ascii_bar(value, max_value, width=30):
    """Generate an ASCII bar for visualizing values in the console.

    Because sometimes simple is better. Like your ex."""

    fill = int((value / max_value) * width) if max_value > 0 else 0

    if USE_COLORS:
        # Choose a color based on the fill percentage

        if fill < width * 0.25:
            color = Fore.RED

        elif fill < width * 0.5:
            color = Fore.YELLOW

        elif fill < width * 0.75:
            color = Fore.CYAN

        else:
            color = Fore.GREEN

        bar = "[" + color + "#" * fill + Style.RESET_ALL + " " * (width - fill) + "]"

    else:
        bar = "[" + "#" * fill + " " * (width - fill) + "]"

    return bar


def generate_report(stats, detailed, word_info, metrics):
    """Generate and print a comprehensive analysis report.

    We don't just report, we BRAG."""

    total = sum(stats.values())

    border = "=" * 78

    print("\n" + border)

    title = "WORD SLAP ULTRA ANALYSIS REPORT"

    if USE_COLORS:
        print(f"{Fore.YELLOW}{title:^78}{Style.RESET_ALL}")

    else:
        print(f"{title:^78}")

    print(border)

    # Add sassy system info

    print(f"\n📊 Report generated on {datetime.now().strftime('%Y-%m-%d at %H:%M:%S')}")

    print(f"🖥️  {platform.system()} {platform.release()} ({platform.machine()})")

    # Summary statistics with ATTITUDE

    print(
        f"\n{Fore.CYAN if USE_COLORS else ''}{'SUMMARY STATISTICS - THE NUMBERS DON'T LIE':^78}{Style.RESET_ALL if USE_COLORS else ''}"
    )

    print("-" * 78)

    # Format metrics in a table with sass

    summary_data = [
        [
            "Total Words Processed",
            f"{metrics['total_words']:,}",
            get_sassy_comment("total_words", metrics["total_words"]),
        ],
        [
            "Unique Words Found",
            f"{metrics['unique_words']:,}",
            get_sassy_comment("unique_words", metrics["unique_words"]),
        ],
        [
            "Unknown Words",
            f"{metrics['unknown_words']:,} ({metrics['unknown_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)",
            "Even I don't know everything!",
        ],
        [
            "Ambiguous Words",
            f"{metrics['ambiguous_words']:,} ({metrics['ambiguous_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)",
            "These words are as confusing as your relationship status!",
        ],
        [
            "Premium-Grade Words",
            f"{metrics['premium_words']:,} ({metrics['premium_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)",
            f"Longer than {PREMIUM_WORD_LENGTH} letters - impressive!",
        ],
        [
            "Average Word Length",
            f"{metrics['avg_word_length']:.2f} characters",
            get_sassy_comment("avg_word_length", metrics["avg_word_length"]),
        ],
        ["Average Syllables", f"{metrics['avg_syllables']:.2f}", "That's a mouthful!"],
        [
            "Average Confidence",
            f"{metrics['avg_confidence']:.2f}",
            "More confident than that person at karaoke night!",
        ],
        [
            "Readability Level",
            f"{metrics['readability_index']}",
            get_sassy_comment("readability_index", metrics["readability_index"]),
        ],
        [
            "Processing Time",
            f"{metrics['processing_time']:.2f} seconds",
            get_sassy_comment("processing_time", metrics["processing_time"]),
        ],
    ]

    if ENABLE_MEMORY_TRACKING:
        summary_data.append(
            [
                "Memory Usage",
                f"{metrics['memory_usage']:.2f} MB",
                get_sassy_comment("memory_usage", metrics["memory_usage"]),
            ]
        )

    if ENABLE_SENTIMENT:
        summary_data.extend(
            [
                [
                    "Positive Words",
                    f"{metrics['positive_words']:,} ({metrics['positive_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)",
                    "Happier than a puppy with two tails!",
                ],
                [
                    "Negative Words",
                    f"{metrics['negative_words']:,} ({metrics['negative_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)",
                    "Darker than your coffee!",
                ],
                [
                    "Neutral Words",
                    f"{metrics['neutral_words']:,} ({metrics['neutral_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)",
                    "As emotional as a brick wall!",
                ],
            ]
        )

    if ENABLE_LINGUISTIC_FEATURES:
        summary_data.extend(
            [
                [
                    "Palindromes",
                    f"{metrics['palindromes']:,}",
                    "Words that read the same backward - how show-offy!",
                ],
                [
                    "Words with Rare Letters",
                    f"{metrics['rare_letter_words']:,}",
                    "Featuring J, K, Q, X, or Z - the cool kids of the alphabet!",
                ],
                [
                    "Words with Prefixes",
                    f"{metrics['prefix_words']:,}",
                    "Prefixed and dangerous!",
                ],
                [
                    "Words with Suffixes",
                    f"{metrics['suffix_words']:,}",
                    "Suffix-icated to perfection!",
                ],
                [
                    "Euphonious Words",
                    f"{metrics['euphonious_words']:,}",
                    "Words that sound as good as chocolate tastes!",
                ],
                [
                    "Alliteration-Ready",
                    f"{metrics['alliteration_words']:,}",
                    "Perfect for tongue twisters and poetry slams!",
                ],
            ]
        )

    # Format and print the table with color

    col_widths = [25, 15, 38]

    # Print table header

    header = ["Metric", "Value", "Commentary"]

    header_fmt = (
        "│ " + " │ ".join(f"{h:{w}}" for h, w in zip(header, col_widths)) + " │"
    )

    if USE_COLORS:
        print(f"{Fore.GREEN}{header_fmt}{Style.RESET_ALL}")

        print(
            "├"
            + "─" * (col_widths[0] + 2)
            + "┼"
            + "─" * (col_widths[1] + 2)
            + "┼"
            + "─" * (col_widths[2] + 2)
            + "┤"
        )

    else:
        print(header_fmt)

        print(
            "│"
            + "-" * (col_widths[0] + 2)
            + "│"
            + "-" * (col_widths[1] + 2)
            + "│"
            + "-" * (col_widths[2] + 2)
            + "│"
        )

    # Print table rows

    for row in summary_data:
        row_fmt = (
            "│ "
            + " │ ".join(f"{str(cell):{w}}" for cell, w in zip(row, col_widths))
            + " │"
        )

        if USE_COLORS:
            print(f"{Fore.CYAN}{row_fmt}{Style.RESET_ALL}")

        else:
            print(row_fmt)

    # Category distribution with bars

    print(
        f"\n{Fore.CYAN if USE_COLORS else ''}{'DISTRIBUTION BY PART OF SPEECH - THE BREAKDOWN':^78}{Style.RESET_ALL if USE_COLORS else ''}"
    )

    print("-" * 78)

    max_count = max(stats.values()) if stats else 0

    table_data = []

    for category in sorted(stats, key=lambda k: -stats[k]):
        count = stats[category]

        percent = (count / total) * 100 if total > 0 else 0

        bar = generate_ascii_bar(count, max_count)

        table_data.append([category, f"{count:,}", f"{percent:.1f}%", bar])

    # Print POS distribution table

    if USE_COLORS:
        print(
            tabulate(
                table_data,
                headers=["Category", "Count", "Percent", "Distribution"],
                tablefmt="pretty",
                colalign=("left", "right", "right", "left"),
            )
        )

    else:
        for cat, count, percent, bar in table_data:
            print(f"{cat:12}: {count:8} words ({percent:5}) {bar}")

    # Word examples with SWAGGER

    print(
        f"\n{Fore.CYAN if USE_COLORS else ''}{'WORD EXAMPLES BY CATEGORY - THE HALL OF FAME':^78}{Style.RESET_ALL if USE_COLORS else ''}"
    )

    print("-" * 78)

    for cat, words in detailed.items():
        if not words:
            continue

        # Sort words by length

        words_by_length = sorted(words, key=len)

        # Sort words by interesting properties

        if len(words) >= 5:
            shortest_words = words_by_length[:3]

            longest_words = words_by_length[-3:]

            # Find words with interesting properties

            interesting_words = []

            sentiment_words = []

            for word in random.sample(words, min(20, len(words))):
                if word in word_info:
                    info = word_info[word]

                    # Check for interesting linguistic features

                    features = info.get("linguistic_features", {})

                    if (
                        features.get("palindrome")
                        or features.get("rare_letter")
                        or features.get("euphony")
                    ):
                        interesting_words.append(word)

                    # Check for strong sentiment

                    sentiment = info.get("sentiment", {})

                    if abs(sentiment.get("compound", 0)) > 0.5:
                        sentiment_words.append(word)

            # Print category header with color

            if USE_COLORS:
                cat_color = COLORS.get(cat, Fore.WHITE)

                print(f"\n{cat_color}{cat}{Style.RESET_ALL} — Total: {len(words):,}")

            else:
                print(f"\n{cat} — Total: {len(words):,}")

            print(f"  Shortest: {', '.join(shortest_words)}")

            print(f"  Longest:  {', '.join(longest_words)}")

            if interesting_words:
                print(f"  Most Interesting: {', '.join(interesting_words[:5])}")

            if sentiment_words and ENABLE_SENTIMENT:
                print(f"  Most Emotional: {', '.join(sentiment_words[:5])}")

            # Most common word lengths in this category

            lengths = [len(w) for w in words]

            common_lengths = Counter(lengths).most_common(3)

            print(
                f"  Most common lengths: "
                + ", ".join([f"{l} chars ({c} words)" for l, c in common_lengths])
            )

        else:
            if USE_COLORS:
                cat_color = COLORS.get(cat, Fore.WHITE)

                print(f"\n{cat_color}{cat}{Style.RESET_ALL} — Total: {len(words):,}")

            else:
                print(f"\n{cat} — Total: {len(words):,}")

            print(f"  All words: {', '.join(words)}")

    # Add some final sassy comments

    sassy_conclusions = [
        "Analysis complete! Your word list has been thoroughly judged and found... interesting.",
        "All done! If words could talk, yours would be thanking me for the attention.",
        "Analysis finished! Your vocabulary has been weighed, measured, and... not found wanting!",
        "Finished! I've seen worse word collections. Not many, but some.",
        "Done! Your words have been analyzed with more precision than a surgeon with OCD.",
        "Analysis complete! These words now know more about themselves than they ever wanted to.",
    ]

    print(f"\n{border}")

    if USE_COLORS:
        print(f"{Fore.YELLOW}{random.choice(sassy_conclusions):^78}{Style.RESET_ALL}")

    else:
        print(f"{random.choice(sassy_conclusions):^78}")

    print(f"{border}")


def create_visualizations(stats, detailed, word_info, metrics, output_dir):
    """Create visual charts of word distribution that will make data scientists weep with joy."""

    try:
        # Set a stylish theme for all charts
        plt.style.use("seaborn-v0_8-darkgrid")

        # CHART 1: Distribution by Part of Speech (Pie Chart)
        plt.figure(figsize=(12, 10))

        # Pie chart for word types
        plt.subplot(2, 1, 1)
        labels = list(stats.keys())
        sizes = list(stats.values())
        colors = [COLORS.get(label, "#888888") for label in labels]
        explode = [
            0.1 if label == max(stats, key=stats.get) else 0 for label in labels
        ]  # Explode largest slice

        wedges, texts, autotexts = plt.pie(
            sizes,
            labels=labels,
            colors=colors,
            autopct="%1.1f%%",
            startangle=90,
            shadow=True,
            explode=explode,
        )

        # Enhance text visibility
        for text in texts:
            text.set_fontsize(12)

        for autotext in autotexts:
            autotext.set_fontsize(12)
            autotext.set_fontweight("bold")
            autotext.set_color("white")

        plt.axis("equal")
        plt.title("Distribution by Part of Speech", fontsize=16, fontweight="bold")

        # Bar chart for word lengths
        plt.subplot(2, 1, 2)
        lengths = sorted(metrics["length_distribution"].keys())
        counts = [metrics["length_distribution"][l] for l in lengths]

        # Get average line position
        avg_length = metrics["avg_word_length"]
        bars = plt.bar(lengths, counts, color="skyblue", edgecolor="navy", alpha=0.7)

        plt.axvline(
            x=avg_length,
            color="crimson",
            linestyle="--",
            linewidth=2,
            label=f"Avg Length: {avg_length:.2f}",
        )

        # Highlight the largest bar
        max_idx = counts.index(max(counts))
        bars[max_idx].set_color("gold")
        bars[max_idx].set_edgecolor("darkgoldenrod")

        plt.xlabel("Word Length (characters)", fontsize=12)
        plt.ylabel("Number of Words", fontsize=12)
        plt.title("Word Length Distribution", fontsize=16, fontweight="bold")
        plt.grid(axis="y", linestyle="--", alpha=0.7)
        plt.legend()
        plt.tight_layout()
        plt.savefig(MAIN_CHART_OUTPUT, dpi=300)
        logging.info(f"Main visualization saved to {MAIN_CHART_OUTPUT}")

        # CHART 2: Detailed POS Distribution (Horizontal Bar Chart)
        plt.figure(figsize=(12, 8))
        categories = list(stats.keys())
        counts = [stats[cat] for cat in categories]

        # Sort by count
        sorted_data = sorted(zip(categories, counts), key=lambda x: x[1])
        categories = [x[0] for x in sorted_data]
        counts = [x[1] for x in sorted_data]

        # Create color map
        colors = [COLORS.get(cat, "#888888") for cat in categories]
        y_pos = np.arange(len(categories))
        bars = plt.barh(y_pos, counts, color=colors, edgecolor="black", alpha=0.8)

        # Add count labels to bars
        for i, bar in enumerate(bars):
            width = bar.get_width()
            label_position = width + (width * 0.01)
            plt.text(
                label_position,
                bar.get_y() + bar.get_height() / 2,
                f"{counts[i]} ({counts[i]/sum(counts)*100:.1f}%)",
                va="center",
                fontweight="bold",
            )

        plt.yticks(y_pos, categories, fontsize=12)
        plt.xlabel("Number of Words", fontsize=12)
        plt.title("Part of Speech Distribution", fontsize=16, fontweight="bold")
        plt.grid(axis="x", linestyle="--", alpha=0.7)
        plt.tight_layout()
        plt.savefig(POS_CHART_OUTPUT, dpi=300)
        logging.info(f"POS visualization saved to {POS_CHART_OUTPUT}")

        # CHART 3: Word Length Histogram (with Density Plot)
        plt.figure(figsize=(12, 8))

        # Get word length data
        lengths = []
        for word in word_info:
            lengths.append(word_info[word]["length"])

        # Create KDE plot
        sns.histplot(
            lengths,
            kde=True,
            color="skyblue",
            edgecolor="navy",
            bins=range(min(lengths), max(lengths) + 2),
            alpha=0.7,
        )

        plt.axvline(
            x=avg_length,
            color="crimson",
            linestyle="--",
            linewidth=2,
            label=f"Avg Length: {avg_length:.2f}",
        )

        plt.xlabel("Word Length (characters)", fontsize=12)
        plt.ylabel("Frequency", fontsize=12)
        plt.title(
            "Word Length Distribution with Density Curve",
            fontsize=16,
            fontweight="bold",
        )
        plt.grid(True, linestyle="--", alpha=0.7)
        plt.legend()
        plt.tight_layout()
        plt.savefig(LENGTH_CHART_OUTPUT, dpi=300)
        logging.info(f"Length visualization saved to {LENGTH_CHART_OUTPUT}")

        # CHART 4: Sentiment Analysis (if enabled)
        if ENABLE_SENTIMENT:
            plt.figure(figsize=(14, 10))

            # Create sentiment data
            sentiments = {
                "Positive": metrics["positive_words"],
                "Neutral": metrics["neutral_words"],
                "Negative": metrics["negative_words"],
            }

            # Pie chart
            plt.subplot(2, 1, 1)
            labels = list(sentiments.keys())
            sizes = list(sentiments.values())
            colors = [COLORS.get(label, "#888888") for label in labels]

            plt.pie(
                sizes,
                labels=labels,
                colors=colors,
                autopct="%1.1f%%",
                startangle=90,
                shadow=True,
                textprops={"fontsize": 12, "fontweight": "bold"},
            )

            plt.axis("equal")
            plt.title("Sentiment Distribution", fontsize=16, fontweight="bold")

            # Bar chart with top positive/negative words
            plt.subplot(2, 1, 2)

            # Get top positive and negative words
            pos_words = []
            neg_words = []

            for word, info in word_info.items():
                sentiment = info.get("sentiment", {})
                compound = sentiment.get("compound", 0)

                if compound >= 0.5:
                    pos_words.append((word, compound))
                elif compound <= -0.5:
                    neg_words.append((word, compound))

            # Sort and get top words
            pos_words = sorted(pos_words, key=lambda x: x[1], reverse=True)[:8]
            neg_words = sorted(neg_words, key=lambda x: x[1])[:8]

            # Create combined chart
            all_words = pos_words + neg_words
            words = [w[0] for w in all_words]
            scores = [w[1] for w in all_words]
            colors = ["green" if s > 0 else "red" for s in scores]
            y_pos = np.arange(len(words))
            plt.barh(y_pos, scores, color=colors, edgecolor="black", alpha=0.7)
            plt.yticks(y_pos, words)
            plt.axvline(x=0, color="black", linestyle="-", linewidth=0.5)
            plt.xlabel("Sentiment Score", fontsize=12)
            plt.title("Top Positive and Negative Words", fontsize=16, fontweight="bold")
            plt.grid(axis="x", linestyle="--", alpha=0.7)
            plt.tight_layout()
            plt.savefig(SENTIMENT_CHART_OUTPUT, dpi=300)
            logging.info(f"Sentiment visualization saved to {SENTIMENT_CHART_OUTPUT}")

        # CHART 5: Word Frequency by Length (Bubble Chart)
        plt.figure(figsize=(12, 8))

        # Get frequency by length
        length_freq = metrics["length_distribution"]
        lengths = list(length_freq.keys())
        freqs = [length_freq[l] for l in lengths]

        # Create bubble sizes based on frequency
        sizes = [f * 50 for f in freqs]  # Scale for visibility

        # Create colormap
        cmap = plt.cm.viridis
        normalized_freqs = [
            (f - min(freqs)) / (max(freqs) - min(freqs))
            if max(freqs) != min(freqs)
            else 0.5
            for f in freqs
        ]
        colors = [cmap(nf) for nf in normalized_freqs]

        # Create scatter plot
        plt.scatter(lengths, freqs, s=sizes, c=colors, alpha=0.7, edgecolors="black")

        for i, length in enumerate(lengths):
            if freqs[i] > sum(freqs) / len(freqs):
                plt.annotate(
                    f"{length} chars",
                    xy=(lengths[i], freqs[i]),
                    xytext=(5, 0),
                    textcoords="offset points",
                    fontsize=9,
                    fontweight="bold",
                )

        plt.xlabel("Word Length (characters)", fontsize=12)
        plt.ylabel("Word Count", fontsize=12)
        plt.title("Word Frequency by Length", fontsize=16, fontweight="bold")
        plt.grid(True, linestyle="--", alpha=0.7)
        plt.colorbar(
            plt.cm.ScalarMappable(
                norm=plt.Normalize(min(freqs), max(freqs)), cmap=cmap
            ),
            label="Frequency",
        )
        plt.tight_layout()
        plt.savefig(FREQ_CHART_OUTPUT, dpi=300)
        logging.info(f"Frequency visualization saved to {FREQ_CHART_OUTPUT}")

        # Generate additional charts if we have enough data
        if metrics["unique_words"] > 100 and ENABLE_CLUSTERING:
            try:
                # CHART 6: Word Clusters (using length and syllables)
                plt.figure(figsize=(12, 8))

                # Extract data points
                word_data = []
                for word, info in word_info.items():
                    if random.random() < 0.8:  # Randomly sample to avoid overcrowding
                        word_data.append(
                            {
                                "word": word,
                                "length": info["length"],
                                "syllables": info.get("syllables", 1),
                                "pos": info["main_type"],
                                "confidence": info["confidence"],
                            }
                        )

                if len(word_data) > 10:
                    # Create plot data
                    x = [d["length"] for d in word_data]
                    y = [d["syllables"] for d in word_data]
                    colors = [COLORS.get(d["pos"], "#888888") for d in word_data]
                    labels = [d["word"] for d in word_data]

                    # Create scatter plot
                    plt.figure(figsize=(12, 8))
                    scatter = plt.scatter(
                        x, y, c=colors, alpha=0.7, s=80, edgecolors="black"
                    )

                    # Add labels for some interesting points
                    for i in range(len(word_data)):
                        # Label some interesting outliers
                        if (
                            word_data[i]["length"] > np.percentile(x, 90)
                            or word_data[i]["syllables"] > np.percentile(y, 90)
                            or (
                                word_data[i]["length"] < np.percentile(x, 10)
                                and word_data[i]["syllables"] > np.percentile(y, 75)
                            )
                        ):
                            plt.annotate(
                                labels[i],
                                xy=(x[i], y[i]),
                                xytext=(5, 5),
                                textcoords="offset points",
                                fontsize=9,
                                fontweight="bold",
                            )

                    # Create legend for parts of speech
                    pos_types = set(d["pos"] for d in word_data)
                    legend_elements = [
                        plt.Line2D(
                            [0],
                            [0],
                            marker="o",
                            color="w",
                            label=pos,
                            markerfacecolor=COLORS.get(pos, "#888888"),
                            markersize=10,
                        )
                        for pos in pos_types
                    ]

                    plt.legend(handles=legend_elements, title="Part of Speech")
                    plt.xlabel("Word Length (characters)", fontsize=12)
                    plt.ylabel("Number of Syllables", fontsize=12)
                    plt.title(
                        "Word Clustering by Length and Syllable Count",
                        fontsize=16,
                        fontweight="bold",
                    )
                    plt.grid(True, linestyle="--", alpha=0.7)
                    plt.tight_layout()
                    plt.savefig(CLUSTER_CHART_OUTPUT, dpi=300)
                    logging.info(
                        f"Cluster visualization saved to {CLUSTER_CHART_OUTPUT}"
                    )

            except Exception as e:
                logging.warning(f"Error creating word cluster visualization: {e}")

    except Exception as e:
        logging.error(f"Error creating visualization: {e}")


# ==================================================================================
# DATA EXPORT - Because if you can't export it, did it even happen?
# ==================================================================================


def save_json(stats, detailed, word_info, metrics, output_path):
    """Save analysis results to JSON file.
    JSON so beautiful it belongs in a museum."""

    try:
        # Create a report time stamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Add some sass to the metadata
        metadata = {
            "generator": "WORD SLAP ULTRA - The Ultimate Word Analyzer",
            "version": "2.0.0",
            "tagline": "Making other word analyzers cry since 2022",
            "report_date": timestamp,
            "execution_environment": f"{platform.system()} {platform.release()} ({platform.machine()})",
            "description": "A comprehensive analysis of word characteristics and distributions.",
        }

        data = {
            "metadata": metadata,
            "summary": dict(stats),
            "percentages": {
                k: round((v / sum(stats.values())) * 100, 2)
                if sum(stats.values()) > 0
                else 0
                for k, v in stats.items()
            },
            "metrics": metrics,
            "words_by_type": {k: sorted(v) for k, v in detailed.items()},
            "per_word_data": word_info,
        }

        # Convert any non-serializable objects to strings or basic types
        def clean_for_json(obj):
            if isinstance(obj, dict):
                return {k: clean_for_json(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean_for_json(item) for item in obj]
            elif isinstance(obj, (set, tuple)):
                return [clean_for_json(item) for item in obj]
            elif isinstance(obj, (int, float, str, bool, type(None))):
                return obj
            else:
                return str(obj)

        # Clean data for JSON serialization
        clean_data = clean_for_json(data)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(clean_data, f, indent=2)

        logging.info(
            f"JSON saved to {output_path} (serialized with the finest artisanal indentation)"
        )
        return True

    except Exception as e:
        logging.error(f"Error saving JSON: {e}")
        return False


def save_yaml(stats, detailed, word_info, metrics, output_path):
    """Save analysis results to YAML file.
    For people who hate curly braces but love indentation."""

    try:
        # Create a report time stamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Add some sass to the metadata
        metadata = {
            "generator": "WORD SLAP ULTRA - The Ultimate Word Analyzer",
            "version": "2.0.0",
            "tagline": "YAML so readable it makes your eyes say 'thank you'",
            "report_date": timestamp,
            "execution_environment": f"{platform.system()} {platform.release()} ({platform.machine()})",
            "description": "A comprehensive analysis of word characteristics and distributions.",
        }

        data = {
            "metadata": metadata,
            "summary": dict(stats),
            "percentages": {
                k: round((v / sum(stats.values())) * 100, 2)
                if sum(stats.values()) > 0
                else 0
                for k, v in stats.items()
            },
            "metrics": dict(metrics),  # Convert defaultdicts to regular dicts
            "words_by_type": {k: sorted(v) for k, v in detailed.items()},
            "per_word_data": word_info,
        }

        # Clean data for YAML serialization (similar to JSON cleaning)
        def clean_for_yaml(obj):
            if isinstance(obj, dict):
                return {k: clean_for_yaml(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean_for_yaml(item) for item in obj]
            elif isinstance(obj, (set, tuple)):
                return [clean_for_yaml(item) for item in obj]
            elif isinstance(obj, (int, float, str, bool, type(None))):
                return obj
            elif isinstance(obj, defaultdict):
                return {k: clean_for_yaml(v) for k, v in dict(obj).items()}
            else:
                return str(obj)

        clean_data = clean_for_yaml(data)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            yaml.dump(clean_data, f, sort_keys=False, default_flow_style=False)

        logging.info(f"YAML saved to {output_path} (indented with love and precision)")
        return True

    except Exception as e:
        logging.error(f"Error saving YAML: {e}")
        return False


def save_csv_grouped(word_info, output_path):
    """Save word classification data to CSV file.
    Because sometimes you just need rows and columns in your life."""

    try:
        grouped = defaultdict(list)

        for word, data in word_info.items():
            grouped[data["main_type"]].append((word, data))

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = [
                "Word",
                "Main_Type",
                "Confidence",
                "Length",
                "Synset_Count",
                "Syllables",
                "Positive_Score",
                "Negative_Score",
                "Neutral_Score",
                "Compound_Score",
                "POS_Frequencies",
                "Linguistic_Features",
            ]

            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            # Write each category in alphabetical order
            for pos in sorted(grouped.keys()):
                for word, info in sorted(grouped[pos], key=itemgetter(0)):
                    sentiment = info.get("sentiment", {})

                    writer.writerow(
                        {
                            "Word": word,
                            "Main_Type": pos,
                            "Confidence": f"{info['confidence']:.2f}",
                            "Length": info["length"],
                            "Synset_Count": info["synset_count"],
                            "Syllables": info.get("syllables", 1),
                            "Positive_Score": f"{sentiment.get('positive', 0):.2f}",
                            "Negative_Score": f"{sentiment.get('negative', 0):.2f}",
                            "Neutral_Score": f"{sentiment.get('neutral', 0):.2f}",
                            "Compound_Score": f"{sentiment.get('compound', 0):.2f}",
                            "POS_Frequencies": json.dumps(info["pos_frequencies"]),
                            "Linguistic_Features": json.dumps(
                                info.get("linguistic_features", {})
                            ),
                        }
                    )

        logging.info(
            f"CSV saved to {output_path} (rows and columns that Excel will actually RESPECT)"
        )
        return True

    except Exception as e:
        logging.error(f"Error saving CSV: {e}")
        return False


def save_xlsx(word_info, stats, metrics, output_path):
    """Save word classification data to Excel file with formatting.
    Because some people still think Excel is an actual database."""

    try:
        wb = Workbook()

        # Create Summary sheet
        summary = wb.active
        summary.title = "Summary"

        # Add title with merged cells and larger font
        summary.append(["WORD SLAP ULTRA - COMPREHENSIVE WORD ANALYSIS"])
        summary.merge_cells("A1:I1")
        summary["A1"].font = Font(size=16, bold=True, color="0000FF")
        summary["A1"].alignment = Alignment(horizontal="center")

        # Add a subtitle with attitude
        summary.append(["Because basic word analysis is for basic people."])
        summary.merge_cells("A2:I2")
        summary["A2"].font = Font(size=12, italic=True, color="800080")
        summary["A2"].alignment = Alignment(horizontal="center")

        # Add report date
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary.append([f"Report generated on {timestamp}"])
        summary.merge_cells("A3:I3")
        summary["A3"].font = Font(size=10)
        summary["A3"].alignment = Alignment(horizontal="center")

        # Add decorative border
        thick_border = Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick"),
        )

        for row in range(1, 4):
            for col in range(1, 10):
                cell = summary.cell(row=row, column=col)
                cell.border = thick_border

        # Add metrics with styled section header
        summary.append([])
        summary.append(["SUMMARY METRICS", "", ""])
        summary["A5"].font = Font(size=14, bold=True, color="006400")
        summary.merge_cells("A5:C5")

        # Add key metrics
        row = 6
        summary.append(["Metric", "Value", "Description"])

        for cell in [summary["A6"], summary["B6"], summary["C6"]]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        for key, value, desc in [
            (
                "Total Words",
                metrics["total_words"],
                "Total words processed during analysis",
            ),
            (
                "Unique Words",
                metrics["unique_words"],
                "Number of unique words after removing duplicates",
            ),
            (
                "Unknown Words",
                metrics["unknown_words"],
                "Words not found in WordNet database",
            ),
            (
                "Ambiguous Words",
                metrics["ambiguous_words"],
                "Words with multiple possible parts of speech",
            ),
            (
                "Average Word Length",
                f"{metrics['avg_word_length']:.2f} characters",
                "Mean length of all unique words",
            ),
            (
                "Average Syllables",
                f"{metrics['avg_syllables']:.2f}",
                "Mean syllable count of all unique words",
            ),
            (
                "Average Confidence",
                f"{metrics['avg_confidence']:.2f}",
                "Average confidence in part of speech classification",
            ),
            (
                "Readability Level",
                metrics["readability_index"],
                "Estimated reading level based on word complexity",
            ),
            (
                "Processing Time",
                f"{metrics['processing_time']:.2f} seconds",
                "Total time taken for analysis",
            ),
        ]:
            row += 1
            summary.append([key, value, desc])

            # Add alternating row colors for readability
            if row % 2 == 0:
                for col in range(1, 4):
                    summary.cell(row=row, column=col).fill = PatternFill(
                        start_color="F0F8FF", end_color="F0F8FF", fill_type="solid"
                    )

        # Add distribution table
        summary.append([])
        row += 2
        summary.append(["PART OF SPEECH DISTRIBUTION", "", ""])
        summary.merge_cells(f"A{row}:C{row}")
        summary[f"A{row}"].font = Font(size=14, bold=True, color="006400")
        row += 1
        summary.append(["Part of Speech", "Count", "Percentage"])

        for cell in [summary[f"A{row}"], summary[f"B{row}"], summary[f"C{row}"]]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        # Add rows for each part of speech
        total = sum(stats.values())
        for pos, count in sorted(stats.items(), key=lambda x: -x[1]):
            row += 1
            percent = (count / total) * 100 if total > 0 else 0
            summary.append([pos, count, f"{percent:.1f}%"])

            # Add color coding
            if pos in COLORS:
                hex_color = COLORS[pos].lstrip("#")
                rgb_color = tuple(
                    int(hex_color[i : i + 2], 16) / 255 for i in (0, 2, 4)
                )
                fill = PatternFill(
                    start_color=hex_color, end_color=hex_color, fill_type="solid"
                )
                summary[f"A{row}"].fill = fill

        # Add sentiment data if available
        if ENABLE_SENTIMENT:
            summary.append([])
            row += 2
            summary.append(["SENTIMENT ANALYSIS", "", ""])
            summary.merge_cells(f"A{row}:C{row}")
            summary[f"A{row}"].font = Font(size=14, bold=True, color="006400")
            row += 1
            summary.append(["Sentiment", "Count", "Percentage"])

            for cell in [summary[f"A{row}"], summary[f"B{row}"], summary[f"C{row}"]]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                )

            row += 1
            pos_count = metrics.get("positive_words", 0)
            neg_count = metrics.get("negative_words", 0)
            neu_count = metrics.get("neutral_words", 0)
            sent_total = pos_count + neg_count + neu_count

            for label, count, color in [
                ("Positive", pos_count, "42f56f"),
                ("Neutral", neu_count, "4287f5"),
                ("Negative", neg_count, "f54242"),
            ]:
                percent = (count / sent_total) * 100 if sent_total > 0 else 0
                summary.append([label, count, f"{percent:.1f}%"])

                # Add color coding
                fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                summary[f"A{row}"].fill = fill
                row += 1

        # Create linguistic features section if enabled
        if ENABLE_LINGUISTIC_FEATURES:
            summary.append([])
            row += 1
            summary.append(["LINGUISTIC FEATURES", "", ""])
            summary.merge_cells(f"A{row}:C{row}")
            summary[f"A{row}"].font = Font(size=14, bold=True, color="006400")
            row += 1
            summary.append(["Feature", "Count", "Percentage"])

            for cell in [summary[f"A{row}"], summary[f"B{row}"], summary[f"C{row}"]]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                )

            for label, count in [
                ("Palindromes", metrics.get("palindromes", 0)),
                ("Words with Rare Letters", metrics.get("rare_letter_words", 0)),
                ("Words with Prefixes", metrics.get("prefix_words", 0)),
                ("Words with Suffixes", metrics.get("suffix_words", 0)),
                ("Euphonious Words", metrics.get("euphonious_words", 0)),
                ("Alliteration-Ready Words", metrics.get("alliteration_words", 0)),
                ("Premium Grade Words", metrics.get("premium_words", 0)),
            ]:
                row += 1
                percent = (
                    (count / metrics["unique_words"]) * 100
                    if metrics["unique_words"] > 0
                    else 0
                )
                summary.append([label, count, f"{percent:.1f}%"])

                # Add alternating row colors
                if row % 2 == 0:
                    for col in range(1, 4):
                        summary.cell(row=row, column=col).fill = PatternFill(
                            start_color="F0F8FF", end_color="F0F8FF", fill_type="solid"
                        )

        # Create syllable distribution section
        summary.append([])
        row += 2
        summary.append(["SYLLABLE DISTRIBUTION", "", ""])
        summary.merge_cells(f"A{row}:C{row}")
        summary[f"A{row}"].font = Font(size=14, bold=True, color="006400")
        row += 1
        summary.append(["Syllable Count", "Number of Words", "Percentage"])

        for cell in [summary[f"A{row}"], summary[f"B{row}"], summary[f"C{row}"]]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        syllable_dist = metrics.get("syllable_distribution", {})
        for syllable_count in sorted(syllable_dist.keys()):
            count = syllable_dist[syllable_count]
            row += 1
            percent = (
                (count / metrics["unique_words"]) * 100
                if metrics["unique_words"] > 0
                else 0
            )
            summary.append([syllable_count, count, f"{percent:.1f}%"])

            # Add alternating row colors
            if row % 2 == 0:
                for col in range(1, 4):
                    summary.cell(row=row, column=col).fill = PatternFill(
                        start_color="F0F8FF", end_color="F0F8FF", fill_type="solid"
                    )

        # Add a note about charts
        summary.append([])
        row += 2
        summary.append(
            ['NOTE: Visualizations are available in the "charts" directory.']
        )
        summary.merge_cells(f"A{row}:E{row}")
        summary[f"A{row}"].font = Font(italic=True)

        # Create Words sheet
        words_sheet = wb.create_sheet(title="Word Classification")

        # Add headers with fancy formatting
        headers = [
            "Word",
            "Type",
            "Confidence",
            "Length",
            "Syllables",
            "Synset Count",
            "Sentiment",
            "POS Frequencies",
            "Linguistic Features",
        ]
        words_sheet.append(headers)

        # Format header row
        for col_idx, header in enumerate(headers, 1):
            cell = words_sheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="000080", end_color="000080", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Group by part of speech
        grouped = defaultdict(list)
        for word, data in word_info.items():
            grouped[data["main_type"]].append((word, data))

        # Write data with category headers and sorting
        row = 1
        for pos in sorted(grouped.keys()):
            # Add category header
            row += 1
            words_sheet.append([pos, "", "", "", "", "", "", "", ""])
            words_sheet.merge_cells(f"A{row}:I{row}")
            cell = words_sheet[f"A{row}"]
            cell.font = Font(bold=True, size=12, color="FFFFFF")

            # Add color coding based on part of speech
            if pos in COLORS:
                hex_color = COLORS[pos].lstrip("#")
                fill = PatternFill(
                    start_color=hex_color, end_color=hex_color, fill_type="solid"
                )
                cell.fill = fill
            else:
                cell.fill = PatternFill(
                    start_color="808080", end_color="808080", fill_type="solid"
                )

            # Sort words by length, then alphabetically
            sorted_words = sorted(grouped[pos], key=lambda x: (x[1]["length"], x[0]))

            # Add words
            for word, info in sorted_words:
                row += 1

                # Get sentiment as descriptive label
                sentiment_label = "Neutral"
                sentiment = info.get("sentiment", {})
                compound = sentiment.get("compound", 0)

                if compound >= SENTIMENT_THRESHOLD:
                    sentiment_label = f"Positive ({compound:.2f})"
                elif compound <= -SENTIMENT_THRESHOLD:
                    sentiment_label = f"Negative ({compound:.2f})"

                # Format linguistic features as readable text
                ling_features = info.get("linguistic_features", {})
                feature_text = ", ".join(
                    key for key, value in ling_features.items() if value is True
                )

                words_sheet.append(
                    [
                        word,
                        pos,
                        f"{info['confidence']:.2f}",
                        info["length"],
                        info.get("syllables", 1),
                        info["synset_count"],
                        sentiment_label,
                        str(info["pos_frequencies"]),
                        feature_text,
                    ]
                )

                # Add zebra striping for readability
                if row % 2 == 0:
                    for col in range(1, 10):
                        words_sheet.cell(row=row, column=col).fill = PatternFill(
                            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                        )

        # Add charts sheet
        charts_sheet = wb.create_sheet(title="Charts")
        charts_sheet.append(["WORD ANALYSIS CHARTS"])
        charts_sheet.merge_cells("A1:G1")
        charts_sheet["A1"].font = Font(size=16, bold=True)
        charts_sheet["A1"].alignment = Alignment(horizontal="center")
        charts_sheet.append(
            ["Charts are saved as image files in the 'charts' directory."]
        )
        charts_sheet.merge_cells("A2:G2")
        charts_sheet["A2"].font = Font(italic=True)
        charts_sheet["A2"].alignment = Alignment(horizontal="center")

        # Add basic charts to Excel
        # POS distribution bar chart
        pos_chart = BarChart()
        pos_chart.title = "Part of Speech Distribution"
        pos_chart.style = 10
        pos_chart.type = "col"
        pos_chart.y_axis.title = "Count"
        pos_chart.x_axis.title = "Part of Speech"
        pos_data = Reference(summary, min_col=2, min_row=row - len(stats), max_row=row)
        pos_cats = Reference(
            summary, min_col=1, min_row=row - len(stats) + 1, max_row=row
        )
        pos_chart.add_data(pos_data, titles_from_data=True)
        pos_chart.set_categories(pos_cats)
        charts_sheet.add_chart(pos_chart, "B4")

        # Add explanatory text about the analysis
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append([])
        charts_sheet.append(["ABOUT THIS ANALYSIS"])
        charts_sheet.merge_cells("A12:G12")
        charts_sheet["A12"].font = Font(size=14, bold=True)

        explanation = textwrap.dedent(
            """
        This analysis was performed using WORD SLAP ULTRA, the most advanced word analysis tool known to humankind. 
        It examines words using WordNet for part of speech classification, SentiWordNet for sentiment analysis, 
        and custom algorithms for syllable counting and linguistic feature detection.
        
        The tool identifies patterns in word usage, complexity, and composition that would take mere mortals 
        years to discover. It categorizes words with a confidence score that indicates how certain the 
        classification is, and provides detailed metrics about word length, syllable count, and more.
        
        Why settle for basic word counts when you can SLAP your words into submission with the most 
        comprehensive analysis available? Your buddy's word analyzer is crying in the corner right now.
        """
        ).strip()

        for i, line in enumerate(explanation.split("\n")):
            if line.strip():
                charts_sheet.append([line])
                charts_sheet.merge_cells(f"A{13+i}:G{13+i}")

        # Adjust column widths on all sheets
        for sheet in [summary, words_sheet, charts_sheet]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                sheet.column_dimensions[column].width = min(adjusted_width, 50)

        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

        logging.info(
            f"Excel file saved to {output_path} (with formatting so beautiful it might make you cry)"
        )
        return True

    except Exception as e:
        logging.error(f"Error saving Excel file: {e}")
        return False


def save_text_report(stats, detailed, metrics, output_path):
    """Save a plain text report for those who live in the terminal."""

    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            # Title and border
            border = "=" * 78
            f.write("\n" + border + "\n")
            f.write(f"{'WORD SLAP ULTRA ANALYSIS REPORT':^78}\n")
            f.write(border + "\n\n")

            # Summary statistics
            f.write(f"{'SUMMARY STATISTICS':^78}\n")
            f.write("-" * 78 + "\n")

            # Format metrics
            for key, value in [
                ("Total Words", f"{metrics['total_words']:,}"),
                ("Unique Words", f"{metrics['unique_words']:,}"),
                (
                    "Unknown Words",
                    f"{metrics['unknown_words']:,} ({metrics['unknown_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)",
                ),
                (
                    "Ambiguous Words",
                    f"{metrics['ambiguous_words']:,} ({metrics['ambiguous_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)",
                ),
                ("Average Word Length", f"{metrics['avg_word_length']:.2f} characters"),
                ("Average Syllables", f"{metrics['avg_syllables']:.2f}"),
                ("Average Confidence", f"{metrics['avg_confidence']:.2f}"),
                ("Readability Level", f"{metrics['readability_index']}"),
                ("Processing Time", f"{metrics['processing_time']:.2f} seconds"),
            ]:
                f.write(f"{key:25}: {value}\n")

            if ENABLE_MEMORY_TRACKING:
                f.write(f"{'Memory Usage':25}: {metrics['memory_usage']:.2f} MB\n")

            # Add sentiment metrics if enabled
            if ENABLE_SENTIMENT:
                f.write("\n")
                f.write(f"{'SENTIMENT BREAKDOWN':^78}\n")
                f.write("-" * 78 + "\n")
                f.write(
                    f"{'Positive Words':25}: {metrics['positive_words']:,} ({metrics['positive_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)\n"
                )
                f.write(
                    f"{'Negative Words':25}: {metrics['negative_words']:,} ({metrics['negative_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)\n"
                )
                f.write(
                    f"{'Neutral Words':25}: {metrics['neutral_words']:,} ({metrics['neutral_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%)\n"
                )

            # Add linguistic features if enabled
            if ENABLE_LINGUISTIC_FEATURES:
                f.write("\n")
                f.write(f"{'LINGUISTIC FEATURES':^78}\n")
                f.write("-" * 78 + "\n")
                f.write(f"{'Palindromes':25}: {metrics['palindromes']:,}\n")
                f.write(
                    f"{'Words with Rare Letters':25}: {metrics['rare_letter_words']:,}\n"
                )
                f.write(f"{'Words with Prefixes':25}: {metrics['prefix_words']:,}\n")
                f.write(f"{'Words with Suffixes':25}: {metrics['suffix_words']:,}\n")
                f.write(f"{'Euphonious Words':25}: {metrics['euphonious_words']:,}\n")
                f.write(
                    f"{'Alliteration-Ready Words':25}: {metrics['alliteration_words']:,}\n"
                )
                f.write(f"{'Premium Grade Words':25}: {metrics['premium_words']:,}\n")

            # Category distribution
            f.write("\n")
            f.write(f"{'DISTRIBUTION BY PART OF SPEECH':^78}\n")
            f.write("-" * 78 + "\n")

            total = sum(stats.values())
            max_count = max(stats.values()) if stats else 0

            for category in sorted(stats, key=lambda k: -stats[k]):
                count = stats[category]
                percent = (count / total) * 100 if total > 0 else 0
                bar = "█" * int((count / max_count) * 40) if max_count > 0 else ""
                f.write(f"{category:12}: {count:8,} words ({percent:5.1f}%) {bar}\n")

            # Word examples
            f.write("\n")
            f.write(f"{'WORD EXAMPLES BY CATEGORY':^78}\n")
            f.write("-" * 78 + "\n")

            for cat, words in detailed.items():
                if not words:
                    continue

                # Sort words by length
                words_by_length = sorted(words, key=len)
                f.write(f"\n{cat} — Total: {len(words):,}\n")

                if len(words) > 5:
                    f.write(
                        f"  Shortest: {', '.join(words_by_length[:min(5, len(words_by_length))])}\n"
                    )
                    f.write(
                        f"  Longest:  {', '.join(words_by_length[-min(5, len(words_by_length)):])}\n"
                    )

                    # Most common word lengths in this category
                    lengths = [len(w) for w in words]
                    common_lengths = Counter(lengths).most_common(3)
                    f.write(
                        f"  Most common lengths: "
                        + ", ".join(
                            [f"{l} chars ({c} words)" for l, c in common_lengths]
                        )
                        + "\n"
                    )
                else:
                    f.write(f"  All words: {', '.join(words)}\n")

            # Add syllable distribution
            f.write("\n")
            f.write(f"{'SYLLABLE DISTRIBUTION':^78}\n")
            f.write("-" * 78 + "\n")

            syllable_dist = metrics.get("syllable_distribution", {})
            for count in sorted(syllable_dist.keys()):
                word_count = syllable_dist[count]
                percent = (
                    (word_count / metrics["unique_words"]) * 100
                    if metrics["unique_words"] > 0
                    else 0
                )
                f.write(
                    f"{count} syllable{'s' if count != 1 else '':9}: {word_count:8,} words ({percent:5.1f}%)\n"
                )

            # Include sassy conclusion
            sassy_conclusions = [
                "Analysis complete! Your word list has been thoroughly SLAPPED into submission.",
                "All done! I've analyzed these words harder than your English teacher ever could.",
                "Analysis finished! Your vocabulary has officially been dominated.",
                "Finished! Shakespeare would be jealous of this analysis.",
                "Done! These words now know who's boss (it's me).",
                "Analysis complete! Your words have been weighed, measured, and judged accordingly.",
            ]
            f.write("\n" + border + "\n")
            f.write(f"{random.choice(sassy_conclusions):^78}\n")
            f.write(border + "\n")

            # Add a footer with time stamp
            timestamp = datetime.now().strftime("%Y-%m-%d at %H:%M:%S")
            f.write(f"\nReport generated on {timestamp}\n")
            f.write(
                f"WORD SLAP ULTRA - Making other analyzers look like kindergarten projects since 2022\n"
            )

        logging.info(
            f"Text report saved to {output_path} (for those who prefer reading like it's 1980)"
        )
        return True

    except Exception as e:
        logging.error(f"Error saving text report: {e}")
        return False


def save_markdown_report(stats, detailed, metrics, output_path):
    """Save a markdown report for those who appreciate beautiful documentation."""

    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            # Title and metadata
            timestamp = datetime.now().strftime("%Y-%m-%d at %H:%M:%S")
            f.write("# WORD SLAP ULTRA Analysis Report\n\n")
            f.write(f"*Report generated on {timestamp}*\n\n")
            f.write("*Because basic word analysis is for basic people.*\n\n")

            # System info
            f.write("## System Information\n\n")
            f.write(f"- **OS:** {platform.system()} {platform.release()}\n")
            f.write(f"- **Processor:** {platform.machine()}\n")
            f.write(f"- **Analysis Time:** {metrics['processing_time']:.2f} seconds\n")
            if ENABLE_MEMORY_TRACKING:
                f.write(f"- **Memory Usage:** {metrics['memory_usage']:.2f} MB\n")

            # Summary statistics
            f.write("\n## Summary Statistics\n\n")
            f.write("| Metric | Value |\n")
            f.write("|--------|-------|\n")
            f.write(f"| Total Words | {metrics['total_words']:,} |\n")
            f.write(f"| Unique Words | {metrics['unique_words']:,} |\n")
            f.write(
                f"| Unknown Words | {metrics['unknown_words']:,} ({metrics['unknown_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%) |\n"
            )
            f.write(
                f"| Ambiguous Words | {metrics['ambiguous_words']:,} ({metrics['ambiguous_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}%) |\n"
            )
            f.write(
                f"| Average Word Length | {metrics['avg_word_length']:.2f} characters |\n"
            )
            f.write(f"| Average Syllables | {metrics['avg_syllables']:.2f} |\n")
            f.write(f"| Average Confidence | {metrics['avg_confidence']:.2f} |\n")
            f.write(f"| Readability Level | {metrics['readability_index']} |\n")

            # Add sentiment metrics if enabled
            if ENABLE_SENTIMENT:
                f.write("\n## Sentiment Analysis\n\n")
                f.write("| Sentiment | Count | Percentage |\n")
                f.write("|-----------|-------|------------|\n")
                f.write(
                    f"| Positive | {metrics['positive_words']:,} | {metrics['positive_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )
                f.write(
                    f"| Negative | {metrics['negative_words']:,} | {metrics['negative_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )
                f.write(
                    f"| Neutral | {metrics['neutral_words']:,} | {metrics['neutral_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )

            # Add linguistic features if enabled
            if ENABLE_LINGUISTIC_FEATURES:
                f.write("\n## Linguistic Features\n\n")
                f.write("| Feature | Count | Percentage |\n")
                f.write("|---------|-------|------------|\n")
                f.write(
                    f"| Palindromes | {metrics['palindromes']:,} | {metrics['palindromes']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )
                f.write(
                    f"| Words with Rare Letters | {metrics['rare_letter_words']:,} | {metrics['rare_letter_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )
                f.write(
                    f"| Words with Prefixes | {metrics['prefix_words']:,} | {metrics['prefix_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )
                f.write(
                    f"| Words with Suffixes | {metrics['suffix_words']:,} | {metrics['suffix_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )
                f.write(
                    f"| Euphonious Words | {metrics['euphonious_words']:,} | {metrics['euphonious_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )
                f.write(
                    f"| Alliteration-Ready Words | {metrics['alliteration_words']:,} | {metrics['alliteration_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )
                f.write(
                    f"| Premium Grade Words | {metrics['premium_words']:,} | {metrics['premium_words']/metrics['unique_words']*100 if metrics['unique_words'] else 0:.1f}% |\n"
                )

            # Category distribution
            f.write("\n## Distribution by Part of Speech\n\n")
            f.write("| Part of Speech | Count | Percentage | Distribution |\n")
            f.write("|----------------|-------|------------|-------------|\n")
            
            total = sum(stats.values())
            max_count = max(stats.values()) if stats else 0
            
            for category in sorted(stats, key=lambda k: -stats[k]):
                count = stats[category]
                percent = (count / total) * 100 if total > 0 else 0
                bar = "█" * int((count / max_count) * 20) if max_count > 0 else ""
                f.write(f"| {category} | {count:,} | {percent:.1f}% | {bar} |\n")

            # Word examples
            f.write("\n## Word Examples by Category\n\n")
            
            for cat, words in detailed.items():
                if not words:
                    continue

                # Sort words by length
                words_by_length = sorted(words, key=len)
                f.write(f"### {cat} — Total: {len(words):,}\n\n")
                
                if len(words) > 5:
                    f.write(
                        f"**Shortest:** {', '.join(words_by_length[:min(5, len(words_by_length))])}\n\n"
                    )
                    f.write(
                        f"**Longest:** {', '.join(words_by_length[-min(5, len(words_by_length)):])}\n\n"
                    )

                    # Most common word lengths in this category
                    lengths = [len(w) for w in words]
                    common_lengths = Counter(lengths).most_common(3)
                    f.write(
                        "**Most common lengths:** "
                        + ", ".join(
                            [f"{l} chars ({c} words)" for l, c in common_lengths]
                        )
                        + "\n\n"
                    )
                else:
                    f.write(f"**All words:** {', '.join(words)}\n\n")

            # Add syllable distribution
            f.write("\n## Syllable Distribution\n\n")
            f.write("| Syllable Count | Number of Words | Percentage |\n")
            f.write("|----------------|----------------|------------|\n")
            
            syllable_dist = metrics.get("syllable_distribution", {})
            for count in sorted(syllable_dist.keys()):
                word_count = syllable_dist[count]
                percent = (
                    (word_count / metrics["unique_words"]) * 100
                    if metrics["unique_words"] > 0
                    else 0
                )
                f.write(f"| {count} | {word_count:,} | {percent:.1f}% |\n")

            # Add visualization references
            f.write("\n## Visualizations\n\n")
            f.write(
                "The following visualizations are available in the `charts` directory:\n\n"
            )
            f.write("1. **Word Distribution** - Overview of word types and lengths\n")
            f.write("2. **POS Distribution** - Detailed breakdown by part of speech\n")
            f.write(
                "3. **Length Distribution** - Analysis of word lengths with density curve\n"
            )
            
            if ENABLE_SENTIMENT:
                f.write("4. **Sentiment Distribution** - Analysis of word sentiments\n")
                
            f.write("5. **Word Frequency** - Frequency analysis by length\n")
            
            if ENABLE_CLUSTERING and metrics["unique_words"] > 100:
                f.write(
                    "6. **Word Clusters** - Clustering of words by linguistic properties\n"
                )

            # Include sassy conclusion
            sassy_conclusions = [
                "Analysis complete! Your word list has been thoroughly SLAPPED into submission.",
                "All done! I've analyzed these words harder than your English teacher ever could.",
                "Analysis finished! Your vocabulary has officially been dominated.",
                "Finished! Shakespeare would be jealous of this analysis.",
                "Done! These words now know who's boss (it's me).",
                "Analysis complete! Your words have been weighed, measured, and judged accordingly.",
            ]
            
            f.write(f"\n> *{random.choice(sassy_conclusions)}*\n\n")
            f.write("---\n\n")
            f.write(
                "*WORD SLAP ULTRA - Making other analyzers look like kindergarten projects since 2022*\n"
            )

        # Also generate HTML from markdown
        try:
            with open(output_path, "r", encoding="utf-8") as md_file:
                md_content = md_file.read()
                html_content = markdown.markdown(md_content, extensions=["tables"])

                # Add some CSS for better styling
                html_style = """
                <style>
                    body { font-family: Arial, sans-serif; line-height: 1.6; max-width: 900px; margin: 0 auto; padding: 20px; }
                    h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
                    h2 { color: #2c3e50; border-bottom: 1px solid #ddd; padding-bottom: 5px; margin-top: 30px; }
                    h3 { color: #3498db; }
                    table { border-collapse: collapse; width: 100%; margin: 20px 0; }
                    th, td { padding: 12px; border: 1px solid #ddd; }
                    th { background-color: #3498db; color: white; text-align: left; }
                    tr:nth-child(even) { background-color: #f2f2f2; }
                    code { background-color: #f8f8f8; padding: 2px 4px; border-radius: 4px; }
                    blockquote { background-color: #f9f9f9; border-left: 5px solid #3498db; padding: 10px 20px; margin: 20px 0; }
                    img { max-width: 100%; height: auto; display: block; margin: 20px auto; }
                </style>
                """

                with open(HTML_OUTPUT, "w", encoding="utf-8") as html_file:
                    html_file.write(
                        "<!DOCTYPE html>\n<html>\n<head>\n<meta charset='utf-8'>\n"
                    )
                    html_file.write("<title>WORD SLAP ULTRA Analysis Report</title>\n")
                    html_file.write(html_style)
                    html_file.write("</head>\n<body>\n")
                    html_file.write(html_content)
                    html_file.write("\n</body>\n</html>")

                logging.info(f"HTML report generated at {HTML_OUTPUT}")

        except Exception as e:
            logging.warning(f"Error generating HTML from markdown: {e}")

        logging.info(
            f"Markdown report saved to {output_path} (for the documentation nerds)"
        )
        return True

    except Exception as e:
        logging.error(f"Error saving markdown report: {e}")
        return False


def save_stats_dump(word_info, stats, metrics, detailed, output_path):
    """Save complete statistics for nerds who want to analyze the analysis."""

    try:
        # Create a deep stats object with everything
        deep_stats = {
            "timestamp": datetime.now().isoformat(),
            "summary_stats": dict(stats),
            "metrics": metrics,
            "word_details": word_info,
            "categorized_words": {k: sorted(v) for k, v in detailed.items()},
            "system_info": {
                "platform": platform.system(),
                "release": platform.release(),
                "version": platform.version(),
                "machine": platform.machine(),
                "processor": platform.processor(),
                "python_version": platform.python_version(),
            },
            "configuration": {
                "cache_size": CACHE_SIZE,
                "batch_size": BATCH_SIZE,
                "max_workers": MAX_WORKERS,
                "min_word_length": MIN_WORD_LENGTH,
                "max_word_length": MAX_WORD_LENGTH,
                "min_confidence": MIN_CONFIDENCE,
                "sentiment_threshold": SENTIMENT_THRESHOLD,
                "premium_word_length": PREMIUM_WORD_LENGTH,
                "report_top_n": REPORT_TOP_N,
                "use_parallel": USE_PARALLEL,
                "use_cache": USE_CACHE,
            },
        }

        # Clean non-serializable objects
        def clean_for_json(obj):
            if isinstance(obj, dict):
                return {k: clean_for_json(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean_for_json(item) for item in obj]
            elif isinstance(obj, (set, tuple)):
                return [clean_for_json(item) for item in obj]
            elif isinstance(obj, defaultdict):
                return {k: clean_for_json(v) for k, v in dict(obj).items()}
            elif isinstance(obj, (int, float, str, bool, type(None))):
                return obj
            else:
                return str(obj)

        # Clean data for JSON serialization
        clean_data = clean_for_json(deep_stats)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(clean_data, f, indent=2)

        logging.info(
            f"Complete stats dump saved to {output_path} (for the true data nerds)"
        )
        return True

    except Exception as e:
        logging.error(f"Error saving stats dump: {e}")
        return False


# ==================================================================================
# MAIN PROGRAM - Where the magic happens
# ==================================================================================


def process_input(input_file):
    """Process the input file and return analysis results."""

    words = read_words(input_file)

    if not words:
        logging.error("No words to process. Exiting.")
        return None, None, None, None

    # Process words
    logging.info(f"Starting analysis of {len(words)} words...")
    logging.info(f"Prepare to be AMAZED by the power of WORD SLAP ULTRA!")

    # Analyze!
    stats, detailed, word_info, metrics = analyze_words(words)

    return stats, detailed, word_info, metrics


def generate_outputs(stats, detailed, word_info, metrics):
    """Generate all output files from analysis results."""

    # Save all the different export formats
    status_json = save_json(stats, detailed, word_info, metrics, JSON_OUTPUT)
    status_yaml = save_yaml(stats, detailed, word_info, metrics, YAML_OUTPUT)
    status_csv = save_csv_grouped(word_info, CSV_OUTPUT)
    status_xlsx = save_xlsx(word_info, stats, metrics, XLSX_OUTPUT)
    status_text = save_text_report(stats, detailed, metrics, TEXT_OUTPUT)
    status_markdown = save_markdown_report(stats, detailed, metrics, MARKDOWN_OUTPUT)
    status_stats_dump = save_stats_dump(
        word_info, stats, metrics, detailed, STATS_DUMP_FILE
    )

    # Create visualizations with attitude
    create_visualizations(stats, detailed, word_info, metrics, CHART_DIR)

    # Return overall status
    return all(
        [
            status_json,
            status_yaml,
            status_csv,
            status_xlsx,
            status_text,
            status_markdown,
            status_stats_dump,
        ]
    )


def main():
    """Main program entry point.
    Where WORD SLAP ULTRA flexes its muscles and makes your buddy's analyzer cry."""

    try:
        # Display attitude-filled banner
        display_banner()

        # Setup
        setup_logging()
        setup_nltk()
        setup_directories()

        # Optional sound effect
        play_sound("start")

        # Process input
        stats, detailed, word_info, metrics = process_input(INPUT_FILE)

        if stats is None:
            logging.error("Analysis failed. Check logs for details.")
            return

        # Generate report to console
        generate_report(stats, detailed, word_info, metrics)

        # Generate output files
        success = generate_outputs(stats, detailed, word_info, metrics)

        # Final report with attitude
        border = "=" * 78
        print(f"\n{border}")
        print(f"{'OUTPUT FILES - THE RECEIPTS OF DOMINATION':^78}")
        print(f"{border}")

        # List of outputs with sassy comments
        outputs = [
            (JSON_OUTPUT, "For the JSON junkies"),
            (YAML_OUTPUT, "For the YAML fanatics"),
            (CSV_OUTPUT, "For the spreadsheet enthusiasts"),
            (XLSX_OUTPUT, "For the Excel addicts"),
            (TEXT_OUTPUT, "For the terminal traditionalists"),
            (MARKDOWN_OUTPUT, "For the documentation devotees"),
            (HTML_OUTPUT, "For the browser brigade"),
            (MAIN_CHART_OUTPUT, "The visual masterpiece"),
            (STATS_DUMP_FILE, "For the data scientists"),
            (LOG_FILE, "For the problem detectives"),
        ]

        for path, desc in outputs:
            exists = os.path.exists(path)
            if USE_COLORS:
                status = (
                    f"{Fore.GREEN}✓{Style.RESET_ALL}"
                    if exists
                    else f"{Fore.RED}✗{Style.RESET_ALL}"
                )
            else:
                status = "✓" if exists else "✗"

            rel_path = os.path.relpath(path, os.getcwd())
            print(f"{status} {rel_path:50} {desc}")

        # Final sassy message
        print(f"\n{border}")

        if success:
            message = random.choice(
                [
                    "Analysis complete! Your buddy's analyzer is now officially OBSOLETE!",
                    "BOOM! Word analysis so thorough it should be ILLEGAL!",
                    "Done! This analysis is so good, it deserves its own trophy!",
                    "Analysis complete! The competition has been OBLITERATED!",
                    "Finished! Your buddy's analyzer just called - it's retiring in SHAME!",
                ]
            )
            if USE_COLORS:
                print(f"{Fore.GREEN}{message:^78}{Style.RESET_ALL}")
            else:
                print(f"{message:^78}")
        else:
            message = "Analysis completed with some issues. Check the logs for details."
            if USE_COLORS:
                print(f"{Fore.YELLOW}{message:^78}{Style.RESET_ALL}")
            else:
                print(f"{message:^78}")

        print(f"{border}")

        # Optional completion sound
        play_sound("complete")

    except Exception as e:
        logging.error(f"Error in main program: {e}", exc_info=True)
        play_sound("error")
        if USE_COLORS:
            print(f"\n{Fore.RED}ERROR: {e}{Style.RESET_ALL}")
        else:
            print(f"\nERROR: {e}")
        print("Check the log file for details.")


if __name__ == "__main__":
    main()


# 👋 💥 SLAPPED!
